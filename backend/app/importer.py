"""Excel workbook importer.

Parses the multi-tab "Upload Format.xlsx" style workbook into normalized
holdings + transactions. On import, only *new* transactions (not already
present for the holding, matched by a content hash) are inserted.
"""
from __future__ import annotations

import hashlib
import re
import xml.etree.ElementTree as ET
import zipfile
from collections import defaultdict
from datetime import date, datetime
from io import BytesIO
from typing import Any, Callable

import openpyxl
from sqlalchemy.orm import Session

from .constants import AssetClass, TxnType
from .models import Holding, Profile, Transaction

_ERR_VALUES = {"#value!", "#n/a", "#ref!", "#div/0!", "#name?", "#null!", "#num!", ""}


def _norm(v: Any) -> str:
    if v is None:
        return ""
    return " ".join(str(v).replace("\n", " ").split()).strip()


def _clean_name(v: Any) -> str:
    s = _norm(v)
    return "" if s.lower() in _ERR_VALUES else s


def _to_float(v: Any) -> float | None:
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "")
    if s.lower() in _ERR_VALUES:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _to_date(v: Any) -> date | None:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y", "%d-%b-%Y", "%d-%b-%y"):
        try:
            return datetime.strptime(str(v).strip(), fmt).date()
        except ValueError:
            continue
    return None


def make_dedup_hash(
    asset_class: str, name: str, identifier: str, txn_date: date | None,
    amount: float, quantity: float, txn_type: str,
) -> str:
    key = "|".join([
        asset_class,
        _norm(name).lower(),
        _norm(identifier).lower(),
        txn_date.isoformat() if txn_date else "",
        f"{round(amount or 0.0, 2)}",
        f"{round(quantity or 0.0, 4)}",
        txn_type,
    ])
    return hashlib.sha256(key.encode("utf-8")).hexdigest()


# A parsed row: contributes to a holding and optionally a transaction.
def _rec(asset_class: str, name: str, identifier: str = "", meta: dict | None = None,
         latest_price: float | None = None, cv_add: float | None = None,
         txn: dict | None = None) -> dict:
    return {
        "asset_class": asset_class,
        "name": name,
        "identifier": identifier,
        "meta": meta or {},
        "latest_price": latest_price,
        "cv_add": cv_add,
        "txn": txn,
    }


def _local(tag: str) -> str:
    return tag.split("}")[-1]


def extract_linked_names(content: bytes, sheet_name: str, column: str = "B") -> dict[int, str]:
    """Extract Excel "linked data type" (Stocks/Geography) display names.

    Cells using Excel's Stocks data type store only a ``#VALUE!`` cache and a
    ``vm`` (value-metadata) reference; the human-readable name lives in the
    workbook's rich-data parts. This resolves, for the given ``column`` of
    ``sheet_name``, a mapping of ``{excel_row_number: display_name}``.

    Returns an empty dict for ordinary workbooks (no rich data) or on any error.
    """
    try:
        with zipfile.ZipFile(BytesIO(content)) as z:
            names = set(z.namelist())
            if "xl/richData/rdrichvalue.xml" not in names or "xl/metadata.xml" not in names:
                return {}

            # Rich-value structures: ordered field-key names per structure.
            structs: list[list[str]] = []
            root = ET.fromstring(z.read("xl/richData/rdrichvaluestructure.xml"))
            for s in root:
                if _local(s.tag) == "s":
                    structs.append([k.get("n") for k in s if _local(k.tag) == "k"])

            # Rich values: (structure index, [field values]).
            rvs: list[tuple[int, list[str]]] = []
            root = ET.fromstring(z.read("xl/richData/rdrichvalue.xml"))
            for rv in root:
                if _local(rv.tag) != "rv":
                    continue
                vals = [(v.text or "") for v in rv if _local(v.tag) == "v"]
                rvs.append((int(rv.get("s", 0)), vals))

            def resolve_name(rvi: int, depth: int = 0) -> str:
                if depth > 5 or rvi < 0 or rvi >= len(rvs):
                    return ""
                s_idx, vals = rvs[rvi]
                keys = structs[s_idx] if s_idx < len(structs) else []
                d = {keys[i]: vals[i] for i in range(min(len(keys), len(vals)))}
                if d.get("%cvi"):  # _linkedentity wrapper -> follow to the core
                    return resolve_name(int(d["%cvi"]), depth + 1)
                for pref in ("_DisplayString", "Name", "Official name", "Ticker symbol"):
                    if d.get(pref):
                        return d[pref]
                return ""

            # metadata.xml: futureMetadata[XLRICHVALUE] -> rich-value indices,
            # and valueMetadata -> index into that future-metadata list.
            future_rvb: list[int] = []
            value_meta: list[int] = []
            meta = ET.fromstring(z.read("xl/metadata.xml"))
            for el in meta:
                ln = _local(el.tag)
                if ln == "futureMetadata" and el.get("name") == "XLRICHVALUE":
                    for bk in el:
                        for node in bk.iter():
                            if _local(node.tag) == "rvb":
                                future_rvb.append(int(node.get("i")))
                elif ln == "valueMetadata":
                    for bk in el:
                        rc = [c for c in bk if _local(c.tag) == "rc"]
                        value_meta.append(int(rc[0].get("v")) if rc else -1)

            def name_for_vm(vm: int) -> str:
                idx = vm - 1
                if not (0 <= idx < len(value_meta)):
                    return ""
                fut = value_meta[idx]
                if not (0 <= fut < len(future_rvb)):
                    return ""
                return resolve_name(future_rvb[fut])

            # Locate the target sheet's XML part.
            wb_xml = z.read("xl/workbook.xml").decode("utf-8", "replace")
            m = re.search(rf'<sheet name="{re.escape(sheet_name)}"[^>]*r:id="(rId\d+)"', wb_xml)
            if not m:
                return {}
            rid = m.group(1)
            rels = z.read("xl/_rels/workbook.xml.rels").decode("utf-8", "replace")
            tm = re.search(rf'Id="{rid}"[^>]*Target="([^"]+)"', rels)
            if not tm:
                return {}
            target = tm.group(1)
            if not target.startswith("xl/"):
                target = "xl/" + target.lstrip("/")

            sheet_xml = z.read(target).decode("utf-8", "replace")
            result: dict[int, str] = {}
            pattern = rf'<c r="{re.escape(column)}(\d+)"[^>]*\bvm="(\d+)"'
            for row_s, vm_s in re.findall(pattern, sheet_xml):
                nm = name_for_vm(int(vm_s))
                if nm:
                    result[int(row_s)] = nm
            return result
    except Exception:
        return {}


def _rows(ws, header_row: int):
    data = list(ws.iter_rows(values_only=True))
    return data[header_row + 1:]


def _cell(row: tuple, idx: int):
    return row[idx] if idx < len(row) else None


# ----------------------- per-sheet parsers -----------------------
def parse_stocks(ws, linked_names: dict[int, str] | None = None) -> list[dict]:
    """Parse the Stocks tab.

    The "Script" column (B) uses Excel's linked "Stocks" data type, so its real
    display name (e.g. "APOLLO HOSPITALS ENTERPRISE LIMITED (XNSE:APOLLOHOSP)")
    is resolved from the workbook's rich data and passed in via ``linked_names``
    (keyed by Excel row number). A few rows use plain text instead (e.g. SGBs);
    those are read directly. As a last-resort fallback (workbooks with neither),
    consecutive rows are grouped into a holding by their sector / market-cap /
    latest-price signature and given a readable label.
    """
    linked_names = linked_names or {}
    parsed = []
    # Header is at 0-based index 1 (Excel row 2), so the k-th data row here is
    # Excel row (k + 3) — used to look up linked-entity names.
    for k, r in enumerate(_rows(ws, 1)):
        excel_row = k + 3
        d = _to_date(_cell(r, 2))
        amount = _to_float(_cell(r, 6))
        sector = _norm(_cell(r, 18))
        # Real purchase rows always carry a sector; rows without one are
        # summary / totals rows in the sheet, so skip them.
        if d is None or amount is None or not sector:
            continue
        parsed.append({
            "real": linked_names.get(excel_row) or _clean_name(_cell(r, 1)),
            "sector": sector,
            "ownership": _norm(_cell(r, 19)),
            "cap_cat": _norm(_cell(r, 20)),
            "cap_cr": _norm(_cell(r, 21)),
            "stop_loss": _norm(_cell(r, 22)),
            "latest": _to_float(_cell(r, 8)),
            "txn": {
                "txn_date": d, "txn_type": TxnType.BUY.value,
                "quantity": _to_float(_cell(r, 3)) or 0.0,
                "price": _to_float(_cell(r, 4)) or 0.0,
                "amount": amount,
                "charges": _to_float(_cell(r, 5)) or 0.0,
                "dividend": _to_float(_cell(r, 11)) or 0.0,
            },
        })

    # Group consecutive rows into stock blocks.
    blocks: list[list[dict]] = []
    prev_sig = object()
    for p in parsed:
        sig = (p["sector"], p["ownership"], p["cap_cat"], p["cap_cr"],
               round(p["latest"], 4) if p["latest"] is not None else None, p["real"])
        if sig != prev_sig:
            blocks.append([])
            prev_sig = sig
        blocks[-1].append(p)

    # Count how many unnamed blocks share the same stable attributes so we only
    # add a numeric suffix when it is actually needed for uniqueness.
    stable_total: dict[tuple, int] = defaultdict(int)
    for b in blocks:
        if not any(x["real"] for x in b):
            stable_total[(b[0]["sector"], b[0]["cap_cat"], b[0]["cap_cr"])] += 1

    out: list[dict] = []
    stable_seen: dict[tuple, int] = defaultdict(int)
    for b in blocks:
        head = b[0]
        real = next((x["real"] for x in b if x["real"]), "")
        stable_key = (head["sector"], head["cap_cat"], head["cap_cr"])
        if real:
            name = real
        else:
            parts = [x for x in (head["sector"], head["cap_cat"], head["cap_cr"])
                     if x and x.upper() != "NA"]
            base = "Stock - " + (" / ".join(parts) if parts else "Unnamed")
            if stable_total[stable_key] > 1:
                stable_seen[stable_key] += 1
                name = f"{base} #{stable_seen[stable_key]}"
            else:
                name = base
        meta = {
            "sector": head["sector"],
            "ownership": head["ownership"],
            "market_cap_category": head["cap_cat"],
            "market_cap_cr": head["cap_cr"],
            "stop_loss": head["stop_loss"],
            "needs_naming": not real,
        }
        for x in b:
            out.append(_rec(
                AssetClass.STOCKS.value, name, meta=meta,
                latest_price=head["latest"], txn=x["txn"],
            ))
    return out


def parse_crypto(ws) -> list[dict]:
    out = []
    for r in _rows(ws, 1):
        name = _clean_name(_cell(r, 2))
        d = _to_date(_cell(r, 3))
        amount = _to_float(_cell(r, 7))
        if not name or d is None or amount is None:
            continue
        out.append(_rec(
            AssetClass.CRYPTO.value, name, identifier=_norm(_cell(r, 1)),
            meta={"exchange": _norm(_cell(r, 1))},
            latest_price=_to_float(_cell(r, 9)),
            txn={
                "txn_date": d, "txn_type": TxnType.BUY.value,
                "quantity": _to_float(_cell(r, 4)) or 0.0,
                "price": _to_float(_cell(r, 5)) or 0.0,
                "amount": amount,
                "charges": _to_float(_cell(r, 6)) or 0.0,
                "dividend": _to_float(_cell(r, 11)) or 0.0,
            },
        ))
    return out


def parse_mf(ws, linked_names: dict[int, str] | None = None) -> list[dict]:
    """Parse the MF tab.

    Like Stocks, the "Scheme Name" column (A) frequently uses Excel's linked
    data type, so the real scheme name is resolved from the workbook rich data
    and passed in via ``linked_names`` (keyed by Excel row number).
    """
    linked_names = linked_names or {}
    out = []
    # Header is at 0-based index 0 (Excel row 1); the k-th data row is Excel row (k + 2).
    for k, r in enumerate(_rows(ws, 0)):
        excel_row = k + 2
        name = linked_names.get(excel_row) or _clean_name(_cell(r, 0))
        d = _to_date(_cell(r, 1))
        amount = _to_float(_cell(r, 2))
        if not name or d is None or amount is None:
            continue
        out.append(_rec(
            AssetClass.MF.value, name,
            meta={"fund_type": _norm(_cell(r, 14)), "comments": _norm(_cell(r, 12))},
            latest_price=_to_float(_cell(r, 5)),
            txn={
                "txn_date": d, "txn_type": TxnType.BUY.value,
                "quantity": _to_float(_cell(r, 4)) or 0.0,
                "price": _to_float(_cell(r, 3)) or 0.0,
                "amount": amount,
                "dividend": _to_float(_cell(r, 7)) or 0.0,
            },
        ))
    return out


def _parse_nps(ws, tier: str, cols: dict[str, int]) -> list[dict]:
    out = []
    for r in _rows(ws, 0):
        name = _clean_name(_cell(r, cols["name"]))
        d = _to_date(_cell(r, cols["date"]))
        amount = _to_float(_cell(r, cols["amount"]))
        if not name or d is None or amount is None:
            continue
        out.append(_rec(
            AssetClass.NPS.value, name, identifier=tier, meta={"tier": tier},
            latest_price=_to_float(_cell(r, cols["latest"])),
            txn={
                "txn_date": d, "txn_type": TxnType.BUY.value,
                "quantity": _to_float(_cell(r, cols["units"])) or 0.0,
                "price": _to_float(_cell(r, cols["nav"])) or 0.0,
                "amount": amount,
                "charges": _to_float(_cell(r, cols["charges"])) or 0.0,
                "dividend": _to_float(_cell(r, cols["div"])) or 0.0,
            },
        ))
    return out


# NPS Tier I has an extra "Shifting Charges" column (index 4), which shifts
# every subsequent column one place to the right relative to Tier II.
def parse_nps_tier1(ws):
    return _parse_nps(ws, "Tier I", {
        "name": 0, "date": 1, "charges": 3, "amount": 5,
        "nav": 6, "units": 7, "latest": 8, "div": 10,
    })


def parse_nps_tier2(ws):
    return _parse_nps(ws, "Tier II", {
        "name": 0, "date": 1, "charges": 3, "amount": 4,
        "nav": 5, "units": 6, "latest": 7, "div": 9,
    })


def parse_fd(ws) -> list[dict]:
    out = []
    for r in _rows(ws, 0):
        acno = _norm(_cell(r, 0))
        principal = _to_float(_cell(r, 1))
        d = _to_date(_cell(r, 7))
        if not acno or principal is None or d is None:
            continue
        meta = {
            "account_no": acno,
            "maturity_date": (_to_date(_cell(r, 8)).isoformat() if _to_date(_cell(r, 8)) else ""),
            "maturity_amount": _to_float(_cell(r, 3)),
            "final_maturity_amount": _to_float(_cell(r, 6)),
            "interest": _to_float(_cell(r, 4)),
            "tds": _to_float(_cell(r, 5)),
            "interest_rate": _to_float(_cell(r, 9)),
        }
        out.append(_rec(
            AssetClass.FD.value, f"FD {acno}", identifier=acno, meta=meta,
            cv_add=_to_float(_cell(r, 2)),  # current accrued value
            txn={"txn_date": d, "txn_type": TxnType.BUY.value, "amount": principal},
        ))
    return out


def parse_pf(ws) -> list[dict]:
    out = []
    for r in _rows(ws, 1):
        d = _to_date(_cell(r, 0))
        total = _to_float(_cell(r, 4))
        company = _norm(_cell(r, 23)) or "PF"
        if d is None or total is None or total == 0:
            continue
        out.append(_rec(
            AssetClass.PF.value, f"PF - {company}", identifier=company,
            meta={"company": company, "interest_rate": _to_float(_cell(r, 5))},
            # Use the compounded latest value ("Compound" column of the Total
            # block, col 21) rather than the simple "Latest Value" (col 19).
            cv_add=_to_float(_cell(r, 21)),
            txn={"txn_date": d, "txn_type": TxnType.BUY.value, "amount": total,
                 "notes": f"Emp {_to_float(_cell(r, 1))}, Employer {_to_float(_cell(r, 2))}, "
                          f"Pension {_to_float(_cell(r, 3))}"},
        ))
    return out


def parse_gratuity(ws) -> list[dict]:
    out = []
    for r in _rows(ws, 0):
        d = _to_date(_cell(r, 0))
        amount = _to_float(_cell(r, 1))
        company = _norm(_cell(r, 8)) or "Gratuity"
        if d is None or amount is None:
            continue
        out.append(_rec(
            AssetClass.GRATUITY.value, f"Gratuity - {company}", identifier=company,
            meta={"company": company, "interest_rate": _to_float(_cell(r, 2))},
            # Current value = compounded value ("Compound" column, index 6)
            # rather than the simple "Latest Value" (index 4).
            cv_add=_to_float(_cell(r, 6)),
            txn={"txn_date": d, "txn_type": TxnType.BUY.value, "amount": amount},
        ))
    return out


def parse_bonds(ws) -> list[dict]:
    out = []
    for r in _rows(ws, 0):
        name = _clean_name(_cell(r, 0))
        d = _to_date(_cell(r, 1))
        amount = _to_float(_cell(r, 2))
        if not name or d is None or amount is None or amount == 0:
            continue
        out.append(_rec(
            AssetClass.BONDS.value, name,
            cv_add=_to_float(_cell(r, 5)),
            meta={"interest_realized": _to_float(_cell(r, 6))},
            txn={"txn_date": d, "txn_type": TxnType.BUY.value, "amount": amount},
        ))
    return out


def parse_policies(ws) -> list[dict]:
    out = []
    # Block 3 (summary) gives policy number + remaining corpus per scheme.
    summary: dict[str, dict] = {}
    for r in _rows(ws, 0):
        sname = _clean_name(_cell(r, 12))
        if sname and sname.lower() != "total":
            summary[sname] = {
                "policy_number": _norm(_cell(r, 13)),
                "remaining": _to_float(_cell(r, 18)),
            }
    for r in _rows(ws, 0):
        # Premium payments (block 1)
        name = _clean_name(_cell(r, 0))
        d = _to_date(_cell(r, 1))
        amount = _to_float(_cell(r, 2))
        if name and d is not None and amount is not None:
            info = summary.get(name, {})
            out.append(_rec(
                AssetClass.POLICIES.value, name, identifier=info.get("policy_number", ""),
                meta={"policy_number": info.get("policy_number", ""),
                      "remaining": info.get("remaining")},
                # Current value = premium grown at 8.5% PA (column E), summed
                # across all premium rows for the policy.
                cv_add=_to_float(_cell(r, 4)),
                txn={"txn_date": d, "txn_type": TxnType.BUY.value, "amount": amount},
            ))
        # Cashbacks / survival benefits (block 2)
        cname = _clean_name(_cell(r, 6))
        cd = _to_date(_cell(r, 7))
        camount = _to_float(_cell(r, 8))
        if cname and cd is not None and camount:
            out.append(_rec(
                AssetClass.POLICIES.value, cname,
                txn={"txn_date": cd, "txn_type": TxnType.CASHBACK.value, "amount": camount},
            ))
    return out


SHEET_PARSERS: dict[str, Callable] = {
    "Stocks": parse_stocks,
    "Crypto": parse_crypto,
    "MF": parse_mf,
    "NPS - Tier I": parse_nps_tier1,
    "NPS - Tier II": parse_nps_tier2,
    "FD": parse_fd,
    "PF": parse_pf,
    "Gratuity": parse_gratuity,
    "Bonds": parse_bonds,
    "Policies": parse_policies,
}


def parse_workbook(content: bytes) -> list[dict]:
    wb = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)
    stock_names = extract_linked_names(content, "Stocks", "B")
    mf_names = extract_linked_names(content, "MF", "A")
    records: list[dict] = []
    for sheet_name, parser in SHEET_PARSERS.items():
        if sheet_name in wb.sheetnames:
            try:
                if sheet_name == "Stocks":
                    records.extend(parse_stocks(wb[sheet_name], stock_names))
                elif sheet_name == "MF":
                    records.extend(parse_mf(wb[sheet_name], mf_names))
                else:
                    records.extend(parser(wb[sheet_name]))
            except Exception as exc:  # keep importing other sheets
                records.append({"_error": f"{sheet_name}: {exc}"})
    wb.close()
    return records


def import_workbook(db: Session, profile: Profile, content: bytes) -> dict:
    records = parse_workbook(content)
    errors = [r["_error"] for r in records if "_error" in r]
    records = [r for r in records if "_error" not in r]

    # Cache existing holdings for this profile.
    holdings: dict[tuple, Holding] = {}
    for h in db.query(Holding).filter(Holding.profile_id == profile.id).all():
        holdings[(h.asset_class, h.name.lower(), h.identifier.lower())] = h

    cv_sums: dict[int, float] = {}          # holding id (or temp key) -> summed current value
    cv_tracks: dict[tuple, float] = {}      # holding key -> summed cv (pre-id)
    latest_prices: dict[tuple, float] = {}

    summary: dict[str, dict] = {}

    def bump(ac: str, field: str, by: int = 1):
        s = summary.setdefault(ac, {"parsed": 0, "new_transactions": 0,
                                    "duplicates": 0, "holdings_created": 0})
        s[field] += by

    # First pass: ensure holdings exist and aggregate holding-level values.
    for rec in records:
        ac = rec["asset_class"]
        key = (ac, rec["name"].lower(), rec["identifier"].lower())
        bump(ac, "parsed")
        holding = holdings.get(key)
        if holding is None:
            holding = Holding(
                profile_id=profile.id, asset_class=ac, name=rec["name"],
                identifier=rec["identifier"], meta=dict(rec["meta"]),
            )
            db.add(holding)
            db.flush()
            holdings[key] = holding
            bump(ac, "holdings_created")
        else:
            if rec["meta"]:
                merged = dict(holding.meta or {})
                merged.update({k: v for k, v in rec["meta"].items() if v not in (None, "")})
                holding.meta = merged
        if rec["latest_price"] is not None:
            latest_prices[key] = rec["latest_price"]
        if rec["cv_add"] is not None:
            cv_tracks[key] = cv_tracks.get(key, 0.0) + rec["cv_add"]

    # Apply holding-level valuations (file is source of truth).
    for key, price in latest_prices.items():
        holdings[key].latest_price = price
    for key, cv in cv_tracks.items():
        holdings[key].current_value = round(cv, 2)

    # Second pass: insert only new transactions (dedup by content hash).
    for rec in records:
        txn = rec["txn"]
        if not txn:
            continue
        ac = rec["asset_class"]
        key = (ac, rec["name"].lower(), rec["identifier"].lower())
        holding = holdings[key]
        h = make_dedup_hash(ac, rec["name"], rec["identifier"], txn.get("txn_date"),
                            txn.get("amount", 0.0), txn.get("quantity", 0.0),
                            txn.get("txn_type", TxnType.BUY.value))
        exists = (
            db.query(Transaction)
            .filter(Transaction.holding_id == holding.id, Transaction.dedup_hash == h)
            .first()
        )
        if exists:
            bump(ac, "duplicates")
            continue
        db.add(Transaction(
            holding_id=holding.id,
            txn_date=txn["txn_date"],
            txn_type=txn.get("txn_type", TxnType.BUY.value),
            quantity=txn.get("quantity", 0.0) or 0.0,
            price=txn.get("price", 0.0) or 0.0,
            amount=txn.get("amount", 0.0) or 0.0,
            charges=txn.get("charges", 0.0) or 0.0,
            dividend=txn.get("dividend", 0.0) or 0.0,
            notes=txn.get("notes", ""),
            source="excel",
            dedup_hash=h,
        ))
        bump(ac, "new_transactions")

    db.commit()

    totals = {
        "parsed": sum(s["parsed"] for s in summary.values()),
        "new_transactions": sum(s["new_transactions"] for s in summary.values()),
        "duplicates": sum(s["duplicates"] for s in summary.values()),
        "holdings_created": sum(s["holdings_created"] for s in summary.values()),
    }
    return {"totals": totals, "by_asset_class": summary, "errors": errors}
